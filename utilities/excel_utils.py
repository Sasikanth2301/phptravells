import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill


def get_row_count(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return (sheet.max_row)

def get_coluumn_count(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return (sheet.max_column)

def read_data(file,sheetname,row_num,column_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return (sheet.cell(row_num,column_num).value)

def write_data(file,sheetname,row_num,column_num,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row_num,column_num).value = data
    workbook.save(file)



